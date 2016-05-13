import sys
from openpyxl import Workbook, load_workbook
from docx import Document
from pprint import pprint

# define all the worksheet locations
first_row = 2
title_column = 3
cwss_column = 6
base_column = 7
attack_column = 8
environmental_column = 9


# extract the findings into a list of dicts
def getFindings(ws):
    findings = []

    # we start in column C and count from row 2 down until we get None
    for x in xrange(first_row,ws.max_row):
        if (ws.cell(row=x, column=title_column).value != None):
            finding = {
                "title" : ws.cell(row=x, column=title_column).value,
                "cwss" : ws.cell(row=x, column=cwss_column).value,
                "base" : ws.cell(row=x, column=base_column).value,
                "attack" : ws.cell(row=x, column=attack_column).value,
                "environmental" : ws.cell(row=x, column=environmental_column).value
            }
            findings.append(finding)
    return findings


def generateDocx(findings):

    document = Document()
    # generate and populate the table
    for finding in findings:    
        table = addTable(document)
        table.cell(0,0).text = finding['title']
        table.cell(1,1).text = "KPMG CWSS Scores"
        table.cell(1,4).text = "Total KPMG CWSS:"
        table.cell(1,5).text = "Total Shell CWSS:"
        table.cell(1,6).text = "Affected hosts"
        table.cell(2,1).text = "Base finding: " + str(finding['base'])
        table.cell(2,2).text = "Attack surface: " + str(finding['attack'])
        table.cell(2,3).text = "Environmental: " + str(finding['environmental'])
        table.cell(2,4).text = str(finding['cwss'])
        table.cell(3,1).text = "Finding\n"
        table.cell(4,1).text = "Impact\n"
        table.cell(5,1).text = "Recommendation\n"
        document.add_page_break()
    document.save('report.docx')


# inserts a findings table into the document
def addTable(document):
    r = 6
    c = 8
    table = document.add_table(rows=r, cols=c)
    # merge top row
    # header row
    table.cell(0,0).merge(table.cell(0,5))
    # left coloured column
    table.cell(1,0).merge(table.cell(5,0))
    # KPMG CWSS Scores
    table.cell(1,1).merge(table.cell(1,3))
    # affected hosts
    table.cell(1,6).merge(table.cell(2,7))
    table.cell(3,6).merge(table.cell(5,7))
    # finding, impact, recommendation
    table.cell(3,1).merge(table.cell(3,5))
    table.cell(4,1).merge(table.cell(4,5))
    table.cell(5,1).merge(table.cell(5,5))
    return table


#main
if len(sys.argv) != 2:
    print "Usage: python %s risk_register.xlsx" % sys.argv[0]
    sys.exit(0)

risk_register = sys.argv[1]

# openpyxl give a 'warnings.warn("Discarded range with reserved name")'
# it doesn't affect the sheet we're interested in, so we can safely
# ignore it for now
#warnings.simplefilter("ignore")
wb = load_workbook(risk_register, read_only=True, data_only=True)

ws = wb['Engagement Findings']

findings = getFindings(ws)
generateDocx(findings)

